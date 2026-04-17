from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .models import WooOrder
from .settings import SpreadsheetFieldRoute, SpreadsheetSettings


class SpreadsheetRouter:
    def __init__(self, settings: SpreadsheetSettings) -> None:
        self.settings = settings

    def export_order(self, order: WooOrder) -> None:
        if not self.settings.enabled:
            return

        workbook_path = Path(self.settings.workbook_path)
        wb = load_workbook(workbook_path) if workbook_path.exists() else Workbook()

        if self.settings.worksheet_name in wb.sheetnames:
            ws = wb[self.settings.worksheet_name]
        else:
            ws = wb.create_sheet(self.settings.worksheet_name)

        enabled_routes = [r for r in self.settings.field_routes if r.enabled]
        if not enabled_routes:
            wb.save(workbook_path)
            return

        headers = [r.target_column for r in enabled_routes]
        if ws.max_row == 1 and all(cell.value is None for cell in ws[1]):
            ws.append(headers)

        row = [self._extract(order.raw, r.woo_field) for r in enabled_routes]
        ws.append(row)
        wb.save(workbook_path)

    def _extract(self, raw: dict[str, Any], dotted_path: str) -> Any:
        cur: Any = raw
        for part in dotted_path.split("."):
            if isinstance(cur, dict):
                cur = cur.get(part)
            else:
                return None
        return cur

    @staticmethod
    def available_woo_fields() -> list[str]:
        return [
            "id",
            "status",
            "currency",
            "date_created",
            "date_created_gmt",
            "date_paid",
            "total",
            "total_tax",
            "shipping_total",
            "billing.first_name",
            "billing.last_name",
            "billing.email",
            "billing.phone",
            "billing.city",
            "billing.state",
            "billing.country",
            "shipping.first_name",
            "shipping.last_name",
            "shipping.city",
            "shipping.state",
            "shipping.country",
            "payment_method",
            "payment_method_title",
            "customer_id",
        ]
